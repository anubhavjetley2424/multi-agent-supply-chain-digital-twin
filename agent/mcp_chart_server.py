import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import matplotlib.pyplot as plt
import os

class VetlabExcelTool:
    def format_csm_exceptions(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb['CSM Tracker']
        
        # Define "Red Flag" style for late materials
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        white_text = Font(color="FFFFFF", bold=True)

        for row in ws.iter_rows(min_row=2):
            expected_date = row[4].value # Arrival Column
            start_date = row[5].value    # Production Column
            
            if expected_date > start_date:
                for cell in row:
                    cell.fill = red_fill
                    cell.font = white_text
        
        wb.save(file_path)

    def create_pricing_chart(self, sku_data):
        # Generates a PNG of Price Impact for the Finance Agent
        skus = [d['sku'] for d in sku_data]
        impact = [d['margin_change'] for d in sku_data]
        
        plt.figure(figsize=(10, 6))
        plt.bar(skus, impact, color='skyblue')
        plt.title('Gross Margin Impact per SKU')
        plt.ylabel('Percentage %')
        
        chart_path = "pricing_impact.png"
        plt.savefig(chart_path)
        plt.close()
        return chart_path