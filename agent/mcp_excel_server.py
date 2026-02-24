from mcp.server.fastmcp import FastMCP
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image as OpenpyxlImage
import os

# Initialize FastMCP Server
mcp = FastMCP("Vetlab-Excel-Manager")

@mcp.tool()
def create_csm_exception_report(data: list, filename: str = "CSM_Exceptions.xlsx"):
    """
    Creates a formatted Excel report for material shortages.
    Applies red highlights to rows where arrival > production start.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CSM Tracker"
    
    # Headers
    headers = ["Order ID", "Material", "Qty", "Expected Arrival", "Production Start"]
    ws.append(headers)
    
    # Styles
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)

    for entry in data:
        row_data = [
            entry['order_id'], entry['material_name'], entry['qty_required'],
            entry['expected_arrival_date'], entry['production_start_date']
        ]
        ws.append(row_data)
        
        # Logic: Red flag if arrival is later than production start
        if entry['expected_arrival_date'] > entry['production_start_date']:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill
                cell.font = white_bold

    file_path = os.path.join(os.getcwd(), filename)
    wb.save(file_path)
    return f"Report generated: {file_path}"

@mcp.tool()
def create_pricing_support_pack(pricing_data: list, chart_path: str = None):
    """
    Generates the Finance Support Pack with XLOOKUP formulas and embeds a margin chart.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing Analysis"
    
    ws.append(["SKU ID", "Old Price", "Proposed Price", "Margin Impact %"])
    
    for i, p in enumerate(pricing_data, start=2):
        ws.cell(row=i, column=1, value=p['sku_id'])
        ws.cell(row=i, column=2, value=p['old_price'])
        ws.cell(row=i, column=3, value=p['new_price'])
        # Add Excel Formula for Margin Impact
        ws.cell(row=i, column=4, value=f"=(C{i}-B{i})/B{i}")
        ws.cell(row=i, column=4).number_format = '0.00%'

    # Embed Chart if provided by mcp_chart_server
    if chart_path and os.path.exists(chart_path):
        img = OpenpyxlImage(chart_path)
        ws.add_image(img, "F2")

    path = "Vetlab_Pricing_Pack.xlsx"
    wb.save(path)
    return f"Pricing pack saved at {path}"

if __name__ == "__main__":
    mcp.run()